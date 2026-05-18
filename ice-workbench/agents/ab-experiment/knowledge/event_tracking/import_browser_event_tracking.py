#!/usr/bin/env python3
"""
浏览器埋点知识库导入脚本
从Excel导入浏览器OneTrack埋点数据到SQLite数据库

优化点（vs 内容中心event_tracking.db）：
1. 新增pages表：存储页面定义（从"附1：页面定义"sheet导入），替代event_relations中的page关联
2. 新增event_page_refs表：事件与页面的多对多关联，替代event_relations中relation_type='page'
3. event_relations简化为只存metric关联
4. 新增sdk_fields表：存储OneTrack SDK系统属性（独立于common_fields）
5. common_field_groups增加scope字段区分公共属性/信息流属性
6. events表新增incognito_report字段（无痕模式是否上报）
7. event_fields表新增is_common_ref字段（标识是否为公共属性引用行）
8. 新增metadata表：存储数据库元信息（版本、来源、更新时间等）
9. 新增categories表：事件分类定义，替代硬编码
"""

import sqlite3
import openpyxl
import re
import sys
import os
from datetime import datetime

EXCEL_PATH = '/Users/cyber_pangda/Downloads/浏览器新版OneTrack埋点汇总（可用于神策、数鲸）.xlsx'
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'browser_event_tracking.db')

CATEGORY_MAP = {
    'app浏览器全局事件': 'app',
    'content信息流事件': 'content',
    'search搜索事件': 'search',
    'livestream直播事件': 'livestream',
    'ad商业化事件': 'ad',
    'personal个人中心事件': 'personal',
    'icon_slots站点事件': 'icon_slots',
    'general常规事件': 'general',
    'setting设置事件': 'setting',
    '信息流热榜内容事件': 'hot_content',
    '工程埋点': 'engineering',
    'novel小说事件': 'novel',
    '热榜事件': 'hot',
    'button_bar底部工具栏事件': 'button_bar',
    'download下载事件': 'download',
    '下载拦截事件': 'download_intercept',
    '浏览器Push事件': 'push',
    'AI搜索': 'ai_search',
    'AI浏览器': 'ai_browser',
    '搜索_安全网址事件(服务端)': 'search_security',
}

SKIP_SHEETS = {
    '说明', 'OneTrack SDK系统属性', 'commen key公共属性',
    'content key信息流通用属性', '附：接入tips', 'bugfix',
    '细节记录', '附1：页面定义', '附2：安全网址请求过滤规则',
    'Sheet30',
}


def create_schema(conn):
    c = conn.cursor()

    c.execute('''CREATE TABLE IF NOT EXISTS metadata (
        key   TEXT PRIMARY KEY,
        value TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS categories (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        name        TEXT NOT NULL UNIQUE,
        name_cn     TEXT,
        description TEXT,
        sort_order  INTEGER DEFAULT 0
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS sdk_fields (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        name        TEXT NOT NULL UNIQUE,
        name_cn     TEXT,
        type        TEXT NOT NULL,
        description TEXT,
        value_note  TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS common_field_groups (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        name        TEXT NOT NULL UNIQUE,
        name_cn     TEXT,
        description TEXT,
        scope       TEXT,
        sort_order  INTEGER DEFAULT 0
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS common_fields (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        group_id    INTEGER NOT NULL,
        classification TEXT,
        name        TEXT NOT NULL,
        name_cn     TEXT,
        type        TEXT NOT NULL,
        description TEXT,
        value_note  TEXT,
        app_ver     TEXT,
        FOREIGN KEY (group_id) REFERENCES common_field_groups(id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS pages (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        page_key    TEXT NOT NULL,
        name_cn     TEXT,
        category    TEXT,
        module_key  TEXT,
        module_cn   TEXT,
        note        TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS events (
        id                INTEGER PRIMARY KEY AUTOINCREMENT,
        name              TEXT NOT NULL UNIQUE,
        name_cn           TEXT NOT NULL,
        category          TEXT,
        classification    TEXT,
        report_when       TEXT,
        description       TEXT,
        note              TEXT,
        incognito_report  TEXT,
        app_ver           TEXT,
        FOREIGN KEY (category) REFERENCES categories(name)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS event_fields (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        event_id      INTEGER NOT NULL,
        name          TEXT NOT NULL,
        name_cn       TEXT,
        type          TEXT,
        description   TEXT,
        value_note    TEXT,
        is_common_ref INTEGER DEFAULT 0,
        common_group  TEXT,
        sort_order    INTEGER DEFAULT 0,
        FOREIGN KEY (event_id) REFERENCES events(id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS event_common_field_refs (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        event_id      INTEGER NOT NULL,
        group_id      INTEGER NOT NULL,
        FOREIGN KEY (event_id) REFERENCES events(id),
        FOREIGN KEY (group_id) REFERENCES common_field_groups(id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS event_page_refs (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        event_id      INTEGER NOT NULL,
        page_id       INTEGER,
        FOREIGN KEY (event_id) REFERENCES events(id),
        FOREIGN KEY (page_id) REFERENCES pages(id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS event_metric_refs (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        event_id      INTEGER NOT NULL,
        metric_name   TEXT NOT NULL,
        FOREIGN KEY (event_id) REFERENCES events(id)
    )''')

    c.execute('CREATE INDEX IF NOT EXISTS idx_events_category ON events(category)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_events_name ON events(name)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_event_fields_event_id ON event_fields(event_id)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_common_fields_group_id ON common_fields(group_id)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_event_common_field_refs_event ON event_common_field_refs(event_id)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_event_page_refs_event ON event_page_refs(event_id)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_event_metric_refs_event ON event_metric_refs(event_id)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_pages_page_key ON pages(page_key)')

    conn.commit()


def import_metadata(conn):
    c = conn.cursor()
    meta = {
        'db_version': '2.0',
        'product': '小米浏览器',
        'source_file': os.path.basename(EXCEL_PATH),
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'appid': '31000000442',
        'hive_table': 'dw.dwd_ot_event_di_31000000442',
        'shujing_url': 'https://s.mi.cn/HCA7W9',
        'onetrack_dashboard': 'https://onetrack.bi.mi.com/#/dashboard?projectId=536',
    }
    for k, v in meta.items():
        c.execute('INSERT OR REPLACE INTO metadata (key, value) VALUES (?, ?)', (k, v))
    conn.commit()


def import_categories(conn):
    c = conn.cursor()
    cats = [
        ('app', '浏览器全局事件', 'App级别全局事件（打开/退出/弹窗/引导/升级等）', 1),
        ('content', '信息流事件', '信息流内容相关事件（曝光/点击/浏览/播放/互动）', 2),
        ('search', '搜索事件', '搜索相关事件', 3),
        ('ad', '商业化事件', '广告商业化事件', 4),
        ('personal', '个人中心事件', '个人中心页面事件', 5),
        ('icon_slots', '站点事件', '名站/宫格/资源位事件', 6),
        ('general', '常规事件', '多窗口/无痕/阅读模式等常规操作事件', 7),
        ('setting', '设置事件', '设置页面事件', 8),
        ('button_bar', '底部工具栏事件', '底部工具栏点击/切换事件', 9),
        ('download', '下载事件', '文件下载相关事件', 10),
        ('download_intercept', '下载拦截事件', '下载拦截相关事件', 11),
        ('novel', '小说事件', '小说功能相关事件', 12),
        ('hot', '热榜事件', '热榜模块事件', 13),
        ('hot_content', '信息流热榜内容事件', '信息流内热榜内容事件', 14),
        ('push', 'Push事件', '浏览器Push推送事件', 15),
        ('livestream', '直播事件', '直播SDK相关事件', 16),
        ('ai_search', 'AI搜索事件', 'AI搜索模块事件', 17),
        ('ai_browser', 'AI浏览器事件', 'AI浏览器功能事件', 18),
        ('engineering', '工程埋点', '研发测试/工程类埋点', 19),
        ('search_security', '安全网址事件', '搜索安全网址服务端事件', 20),
    ]
    for name, name_cn, desc, sort in cats:
        c.execute('INSERT OR IGNORE INTO categories (name, name_cn, description, sort_order) VALUES (?, ?, ?, ?)',
                  (name, name_cn, desc, sort))
    conn.commit()


def import_sdk_fields(conn, wb):
    c = conn.cursor()
    ws = wb['OneTrack SDK系统属性']
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row or not row[1]:
            continue
        name = str(row[1]).strip()
        if not name or name == 'key name\n（English）':
            continue
        name_cn = str(row[2]).strip() if row[2] else ''
        ftype = str(row[3]).strip() if row[3] else ''
        desc = str(row[4]).strip() if row[4] else ''
        c.execute('INSERT OR IGNORE INTO sdk_fields (name, name_cn, type, description, value_note) VALUES (?, ?, ?, ?, ?)',
                  (name, name_cn, ftype, desc, ''))
    conn.commit()
    count = c.execute('SELECT COUNT(*) FROM sdk_fields').fetchone()[0]
    print(f'  SDK系统属性: {count} 条')


def import_common_fields(conn, wb):
    c = conn.cursor()

    # 公共属性组
    groups = [
        ('common_key', '公共属性', '业务定义的公共属性，每个事件都携带', 'all', 1),
        ('content_key', '信息流通用属性', '业务定义的信息流属性，每个信息流事件都携带', 'content', 2),
    ]
    for name, name_cn, desc, scope, sort in groups:
        c.execute('INSERT OR IGNORE INTO common_field_groups (name, name_cn, description, scope, sort_order) VALUES (?, ?, ?, ?, ?)',
                  (name, name_cn, desc, scope, sort))

    # 导入common key
    ws = wb['commen key公共属性']
    group_id = c.execute("SELECT id FROM common_field_groups WHERE name='common_key'").fetchone()[0]
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row or not row[1]:
            continue
        name = str(row[1]).strip()
        if not name or 'key name' in name.lower():
            continue
        classification = str(row[0]).strip() if row[0] else ''
        name_cn = str(row[2]).strip() if row[2] else ''
        ftype = str(row[3]).strip() if row[3] else ''
        value_note = str(row[4]).strip() if row[4] else ''
        desc = str(row[5]).strip() if row[5] else ''
        app_ver = str(row[6]).strip() if len(row) > 6 and row[6] else ''
        c.execute('''INSERT INTO common_fields (group_id, classification, name, name_cn, type, description, value_note, app_ver)
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                  (group_id, classification, name, name_cn, ftype, desc, value_note, app_ver))

    # 导入content key
    ws = wb['content key信息流通用属性']
    group_id = c.execute("SELECT id FROM common_field_groups WHERE name='content_key'").fetchone()[0]
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row or not row[2]:
            continue
        name = str(row[2]).strip()
        if not name or 'key name' in name.lower():
            continue
        classification = str(row[0]).strip() if row[0] else ''
        key_type = str(row[1]).strip() if row[1] else ''
        full_class = f'{classification} | {key_type}' if classification and key_type else (classification or key_type)
        name_cn = str(row[3]).strip() if row[3] else ''
        ftype = str(row[4]).strip() if row[4] else ''
        value_note = str(row[5]).strip() if row[5] else ''
        desc = str(row[6]).strip() if row[6] else ''
        app_ver = str(row[7]).strip() if len(row) > 7 and row[7] else ''
        c.execute('''INSERT INTO common_fields (group_id, classification, name, name_cn, type, description, value_note, app_ver)
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                  (group_id, full_class, name, name_cn, ftype, desc, value_note, app_ver))

    conn.commit()
    count = c.execute('SELECT COUNT(*) FROM common_fields').fetchone()[0]
    print(f'  公共字段: {count} 条')


def import_pages(conn, wb):
    c = conn.cursor()
    ws = wb['附1：页面定义']
    current_category = ''
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row or not row[1]:
            continue
        if row[0]:
            current_category = str(row[0]).strip()
        page_key = str(row[1]).strip()
        name_cn = str(row[2]).strip() if row[2] else ''
        module_key = str(row[3]).strip() if row[3] else ''
        module_cn = str(row[4]).strip() if row[4] else ''
        note = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        if page_key and page_key != 'page / from page\n(英文)':
            c.execute('''INSERT OR IGNORE INTO pages (page_key, name_cn, category, module_key, module_cn, note)
                         VALUES (?, ?, ?, ?, ?, ?)''',
                      (page_key, name_cn, current_category, module_key, module_cn, note))
    conn.commit()
    count = c.execute('SELECT COUNT(*) FROM pages').fetchone()[0]
    print(f'  页面定义: {count} 条')


def _detect_column_layout(ws):
    """检测Excel sheet的列布局，返回列索引映射"""
    row1 = [str(c.value).strip() if c.value else '' for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    row2 = [str(c.value).strip() if c.value else '' for c in list(ws.iter_rows(min_row=2, max_row=2))[0]]

    headers = []
    for i in range(max(len(row1), len(row2))):
        h = row1[i] if row1[i] else row2[i]
        headers.append(h.lower())

    layout = {}
    for i, h in enumerate(headers):
        if 'event name' in h or '事件名（英文）' in h or '事件名(英文)' in h:
            layout['event_name'] = i
        elif '事件名（中文）' in h or '事件名(中文)' in h:
            layout['event_name_cn'] = i
        elif 'report logic' in h or '上报逻辑' in h or '上报时机' in h or 'report time' in h:
            layout['report_logic'] = i
        elif 'key name' in h or '属性名（英文）' in h or '属性名(英文)' in h:
            layout['key_name'] = i
        elif 'key name\n（chinese）' in h or '属性名（中文）' in h or '属性名(中文)' in h:
            layout['key_name_cn'] = i
        elif 'key value type' in h or '属性值类型' in h:
            layout['key_type'] = i
        elif 'key value' in h or '属性名的对应值' in h:
            layout['key_value'] = i
        elif 'remark' in h or '备注' in h:
            if 'remark' not in layout:
                layout['remark'] = i
        elif 'app_ver' in h or '进版版本号' in h:
            layout['app_ver'] = i
        elif '无痕' in h:
            layout['incognito'] = i
        elif 'classification' in h or '分类' in h:
            layout['classification'] = i

    # Fallback: if event_name not found but row2 has 'event name' pattern
    if 'event_name' not in layout:
        for i, h in enumerate(headers):
            if 'event name' in h:
                layout['event_name'] = i
                break

    # Special handling for novel sheet: event name is in col 0 with header '时长'
    if 'event_name' not in layout:
        row2_vals = [str(c.value).strip() if c.value else '' for c in list(ws.iter_rows(min_row=2, max_row=2))[0]]
        for i, h in enumerate(row2_vals):
            if 'event name' in h.lower() or 'event name' in h:
                layout['event_name'] = i
                break

    # If still not found, try to detect from data pattern
    if 'event_name' not in layout:
        for row in ws.iter_rows(min_row=3, max_row=min(10, ws.max_row), values_only=True):
            if not row:
                continue
            for i, cell in enumerate(row):
                if cell and re.match(r'^[a-z][a-z0-9_]+$', str(cell).strip()):
                    if i not in layout.values():
                        layout['event_name'] = i
                        break
            if 'event_name' in layout:
                break

    # Ensure key_name is found
    if 'key_name' not in layout:
        for i, h in enumerate(headers):
            if '属性名' in h:
                layout['key_name'] = i
                break

    return layout


def _is_common_key_ref(val):
    """判断是否为公共属性引用行"""
    if not val:
        return False
    v = str(val).strip().lower().replace('\u200b', '').replace(' ', '')
    return v in ('commonkey', 'commenkey', 'commey_key', 'common_key', 'comment_key',
                 'contentkey', 'content_key')


def _get_common_group(val):
    """根据公共属性引用行判断是common_key还是content_key"""
    if not val:
        return None
    v = str(val).strip().lower().replace('\u200b', '').replace(' ', '')
    if 'content' in v:
        return 'content_key'
    return 'common_key'


def parse_event_rows(ws, category):
    """解析事件sheet，返回事件列表"""
    layout = _detect_column_layout(ws)
    if 'event_name' not in layout or 'key_name' not in layout:
        print(f'    跳过 {category}: 无法检测列布局')
        return []

    events = []
    current_event = None
    event_sort = 0

    start_row = 3
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, values_only=True):
        if not row:
            continue

        row_list = list(row) + [None] * 20
        en = row_list[layout.get('event_name', 0)]
        encn = row_list[layout.get('event_name_cn', 1)]
        report = row_list[layout.get('report_logic', 2)]
        kn = row_list[layout.get('key_name', 3)]
        kncn = row_list[layout.get('key_name_cn', 4)]
        kt = row_list[layout.get('key_type', 5)]
        kv = row_list[layout.get('key_value', 6)]
        remark = row_list[layout.get('remark', 7)]
        app_ver = row_list[layout.get('app_ver', 8)] if 'app_ver' in layout else None
        incognito = row_list[layout.get('incognito', 9)] if 'incognito' in layout else None
        classification = row_list[layout.get('classification', 0)] if 'classification' in layout else None

        en_str = str(en).strip() if en else ''
        encn_str = str(encn).strip() if encn else ''
        kn_str = str(kn).strip() if kn else ''
        kncn_str = str(kncn).strip() if kncn else ''

        # Skip header-like rows
        if en_str and ('event name' in en_str.lower() or '事件名' in en_str):
            continue
        if kn_str and ('key name' in kn_str.lower() or '属性名' in kn_str):
            continue

        # New event
        if en_str and en_str != '':
            if current_event:
                events.append(current_event)
            event_sort += 1
            current_event = {
                'name': en_str,
                'name_cn': encn_str,
                'category': category,
                'classification': str(classification).strip() if classification else '',
                'report_when': str(report).strip() if report else '',
                'note': str(remark).strip() if remark else '',
                'incognito_report': str(incognito).strip() if incognito else '',
                'app_ver': str(app_ver).strip() if app_ver else '',
                'fields': [],
                'common_refs': set(),
                'sort_order': event_sort,
            }
            # Check if the first field in same row is a common key ref
            if kn_str and _is_common_key_ref(kn_str):
                current_event['common_refs'].add(_get_common_group(kn_str))
            elif kn_str and kn_str != '':
                current_event['fields'].append({
                    'name': kn_str,
                    'name_cn': kncn_str,
                    'type': str(kt).strip() if kt else '',
                    'value_note': str(kv).strip() if kv else '',
                    'description': str(remark).strip() if remark else '',
                    'is_common_ref': 0,
                    'common_group': None,
                    'sort_order': len(current_event['fields']),
                })
        elif current_event and kn_str and kn_str != '':
            if _is_common_key_ref(kn_str):
                current_event['common_refs'].add(_get_common_group(kn_str))
            else:
                current_event['fields'].append({
                    'name': kn_str,
                    'name_cn': kncn_str,
                    'type': str(kt).strip() if kt else '',
                    'value_note': str(kv).strip() if kv else '',
                    'description': str(remark).strip() if remark else '',
                    'is_common_ref': 0,
                    'common_group': None,
                    'sort_order': len(current_event['fields']),
                })

    if current_event:
        events.append(current_event)
    return events


def import_events(conn, wb):
    c = conn.cursor()
    total_events = 0
    total_fields = 0
    total_common_refs = 0

    for sheet_name, category in CATEGORY_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f'    跳过 {category}: sheet不存在')
            continue

        ws = wb[sheet_name]
        events = parse_event_rows(ws, category)

        for ev in events:
            # Skip dirty data
            if ev['name'] in ('event name', 'event name\n（English）', '事件名（英文）'):
                continue
            if not ev['name_cn'] and not ev['report_when']:
                continue

            try:
                c.execute('''INSERT INTO events (name, name_cn, category, classification, report_when, description, note, incognito_report, app_ver)
                             VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                          (ev['name'], ev['name_cn'], ev['category'], ev['classification'],
                           ev['report_when'], '', ev['note'], ev['incognito_report'], ev['app_ver']))
                event_id = c.execute('SELECT last_insert_rowid()').fetchone()[0]
                total_events += 1
            except sqlite3.IntegrityError:
                continue

            for f in ev['fields']:
                if f['name'] in ('key name', '属性名（英文）', '属性名(英文)'):
                    continue
                c.execute('''INSERT INTO event_fields (event_id, name, name_cn, type, description, value_note, is_common_ref, common_group, sort_order)
                             VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                          (event_id, f['name'], f['name_cn'], f['type'], f['description'],
                           f['value_note'], f['is_common_ref'], f['common_group'], f['sort_order']))
                total_fields += 1

            for ref in ev['common_refs']:
                group_row = c.execute("SELECT id FROM common_field_groups WHERE name=?", (ref,)).fetchone()
                if group_row:
                    c.execute('INSERT INTO event_common_field_refs (event_id, group_id) VALUES (?, ?)',
                              (event_id, group_row[0]))
                    total_common_refs += 1

    conn.commit()
    print(f'  事件: {total_events} 条')
    print(f'  事件字段: {total_fields} 条')
    print(f'  公共属性引用: {total_common_refs} 条')


def auto_link_pages(conn):
    """根据事件名和page字段自动关联事件与页面"""
    c = conn.cursor()

    page_link_rules = {
        'app_open': ['home', 'gongge'],
        'app_exit': ['home'],
        'app_duration': ['home', 'feed_info_topnews', 'feed_content_detail'],
        'app_background': ['home'],
        'app_foreground': ['home'],
        'content_item_expose': ['feed_info_topnews', 'feed_info_rec', 'feed_video_rec', 'feed_video_shortv'],
        'content_item_click': ['feed_info_topnews', 'feed_info_rec', 'feed_video_rec', 'feed_video_shortv'],
        'content_detail_expose': ['feed_content_detail', 'feed_content_detail_shortvideo'],
        'content_detail_click': ['feed_content_detail'],
        'content_over': ['feed_content_detail', 'feed_content_detail_shortvideo'],
        'content_stay_duration': ['feed_content_detail'],
        'content_comment_expose': ['feed_content_comments_detail'],
        'content_comment_click': ['feed_content_comments_detail'],
        'content_comment_submit': ['feed_content_comments_detail'],
        'content_share': ['feed_content_detail'],
        'content_like': ['feed_content_detail'],
        'content_dislike': ['feed_content_detail'],
        'content_favorite': ['feed_content_detail'],
        'content_unfavorite': ['feed_content_detail'],
        'content_readmode_enter': ['readmode'],
        'content_atlas_enter': ['atlas'],
        'feed_expose': ['feed_info_topnews', 'feed_info_rec', 'feed_video_rec'],
        'feed_click': ['feed_info_topnews', 'feed_info_rec', 'feed_video_rec'],
        'feed_minivideo_expose': ['feed_minivideo_continuously', 'feed_minivideo_continuously_root'],
        'feed_shortvideo_expose': ['feed_shortvideo_immerse', 'feed_shortvideo_immerse_root'],
        'feed_livestream_expose': ['feed_livestream_immersion'],
        'feed_channel_switch': ['feed_info_topnews', 'feed_video_rec'],
        'search': ['search_result'],
        'search_sug_click': ['search_sug'],
        'search_sug_expose': ['search_sug'],
        'search_result_click': ['search_result'],
        'search_home_expose': ['search_home'],
        'search_home_click': ['search_home'],
        'button_bar_click': ['home'],
        'setting_click': ['setting'],
        'personal_center_expose_client': ['me'],
        'personal_center_click': ['me'],
        'novel_expose': ['novel'],
        'novel_read_duration': ['novel'],
        'hot_card_item_expose': ['feed_info_topnews'],
        'hot_card_item_click': ['feed_info_topnews'],
        'hot_list_item_expose': ['feed_info_topnews'],
        'hot_list_item_click': ['feed_info_topnews'],
        'icon_expose': ['home', 'gongge'],
        'icon_click': ['home', 'gongge'],
        'tabaction_expose': ['window_window'],
        'tabaction_click': ['window_window'],
        'download': ['web_page'],
        'push_expose': ['home'],
        'push_click': ['home'],
        'ai_content_expose': ['feed_info_topnews', 'feed_info_rec'],
        'ai_content_click': ['feed_info_topnews', 'feed_info_rec'],
        'AI_homepage_expose': ['home'],
        'AI_homepage_click': ['home'],
        'ad_content_expose': ['feed_info_topnews', 'feed_info_rec', 'feed_video_rec'],
        'ad_content_click': ['feed_info_topnews', 'feed_info_rec', 'feed_video_rec'],
        'bookmark_click': ['bookmark_BookmarkAndHistory'],
        'bookmark_expose': ['bookmark_BookmarkAndHistory'],
        'history_click': ['bookmark_BookmarkAndHistory'],
        'offlinevideo_expose': ['offlinevideo'],
        'qrcode_scan': ['qrcode'],
        'guide_expose': ['guide'],
        'guide_click': ['guide'],
        'feedback_submit': ['feedback'],
        'wallpaper_set': ['wallpaper_CustomWallpaper'],
        'gallery_view': ['gallery'],
        'splash_expose': ['launch_splash'],
        'netdiagno_click': ['netdiagno_NetWorkDiagnostic'],
        'permission_request': ['permission_WebPermissionDetail'],
        'web_page_load': ['web_page'],
        'web_page_error': ['web_page'],
        'app_upgrade': ['home'],
        'app_popup': ['home'],
        'app_guide': ['home', 'guide'],
        'app_notification': ['home'],
        'app_shortcut': ['home'],
        'app_widget': ['home'],
    }

    linked = 0
    for event_name, page_keys in page_link_rules.items():
        event_row = c.execute("SELECT id FROM events WHERE name=?", (event_name,)).fetchone()
        if not event_row:
            continue
        event_id = event_row[0]
        for pk in page_keys:
            page_row = c.execute("SELECT id FROM pages WHERE page_key=?", (pk,)).fetchone()
            if page_row:
                try:
                    c.execute('INSERT OR IGNORE INTO event_page_refs (event_id, page_id) VALUES (?, ?)',
                              (event_id, page_row[0]))
                    linked += 1
                except sqlite3.IntegrityError:
                    pass

    # Also link events that have a 'page' or 'from_page' field to pages
    events_with_page = c.execute("""
        SELECT DISTINCT ef.event_id, ef.value_note
        FROM event_fields ef
        WHERE ef.name IN ('page', 'from_page') AND ef.value_note IS NOT NULL AND ef.value_note != ''
    """).fetchall()

    for event_id, page_values in events_with_page:
        page_keys = re.findall(r'[a-z][a-z0-9_]+', str(page_values))
        for pk in page_keys:
            page_row = c.execute("SELECT id FROM pages WHERE page_key=?", (pk,)).fetchone()
            if page_row:
                try:
                    c.execute('INSERT OR IGNORE INTO event_page_refs (event_id, page_id) VALUES (?, ?)',
                              (event_id, page_row[0]))
                    linked += 1
                except sqlite3.IntegrityError:
                    pass

    # Link events by name pattern to pages
    pattern_rules = [
        (r'^content_', ['feed_content_detail', 'feed_info_topnews']),
        (r'^feed_', ['feed_info_topnews', 'feed_info_rec']),
        (r'^search_', ['search_home', 'search_sug', 'search_result']),
        (r'^novel_', ['novel']),
        (r'^ad_', ['feed_info_topnews', 'feed_info_rec', 'feed_video_rec']),
        (r'^hot_', ['feed_info_topnews']),
        (r'^setting_', ['setting']),
        (r'^personal_', ['me']),
        (r'^download_', ['web_page']),
        (r'^push_', ['home']),
        (r'^icon_', ['home', 'gongge']),
        (r'^tabaction_', ['window_window']),
        (r'^button_bar_', ['home']),
        (r'^ai_', ['feed_info_topnews', 'home']),
        (r'^AI_', ['home']),
        (r'^bookmark_', ['bookmark_BookmarkAndHistory']),
        (r'^livestream_', ['feed_livestream_immersion']),
    ]

    for pattern, page_keys in pattern_rules:
        events = c.execute("SELECT id, name FROM events WHERE name REGEXP ?", (pattern,)).fetchall() if False else []
        events = c.execute("SELECT id, name FROM events").fetchall()
        for event_id, event_name in events:
            if re.match(pattern, event_name):
                for pk in page_keys:
                    page_row = c.execute("SELECT id FROM pages WHERE page_key=?", (pk,)).fetchone()
                    if page_row:
                        try:
                            c.execute('INSERT OR IGNORE INTO event_page_refs (event_id, page_id) VALUES (?, ?)',
                                      (event_id, page_row[0]))
                            linked += 1
                        except sqlite3.IntegrityError:
                            pass

    conn.commit()
    count = c.execute('SELECT COUNT(*) FROM event_page_refs').fetchone()[0]
    print(f'  事件-页面关联: {count} 条')


def auto_link_metrics(conn):
    """自动关联事件与核心指标"""
    c = conn.cursor()

    metric_rules = {
        '浏览器DAU/日活': [
            'app_open',
        ],
        '浏览器信息流DAU/日活': [
            'content_item_expose', 'feed_expose',
            'content_item_click', 'feed_click',
        ],
        '浏览器信息流消费DAU/日活': [
            'content_item_click', 'content_detail_expose',
            'content_over', 'content_stay_duration',
        ],
        '浏览器信息流次日留存率': [
            'content_item_expose', 'feed_expose',
            'content_item_click',
        ],
        '浏览器信息流人均时长': [
            'app_duration', 'content_stay_duration',
            'feed_stay_duration',
        ],
        '浏览器信息流人均消费时长': [
            'content_stay_duration', 'content_over',
            'feed_stay_duration',
        ],
        '浏览器信息流人均VV': [
            'content_over', 'video_play', 'video_over',
            'feed_video_play',
        ],
        '信息流有效用户': [
            'content_item_click', 'content_over',
            'content_stay_duration',
        ],
    }

    linked = 0
    for metric_name, event_names in metric_rules.items():
        for en in event_names:
            event_row = c.execute("SELECT id FROM events WHERE name=?", (en,)).fetchone()
            if event_row:
                try:
                    c.execute('INSERT OR IGNORE INTO event_metric_refs (event_id, metric_name) VALUES (?, ?)',
                              (event_row[0], metric_name))
                    linked += 1
                except sqlite3.IntegrityError:
                    pass

    # Also link by event name pattern
    pattern_metric_rules = [
        (r'^content_item_expose', '浏览器信息流DAU/日活'),
        (r'^content_item_click', '浏览器信息流消费DAU/日活'),
        (r'^content_over', '浏览器信息流人均VV'),
        (r'^content_stay_duration', '浏览器信息流人均消费时长'),
        (r'^app_open', '浏览器DAU/日活'),
        (r'^app_duration', '浏览器信息流人均时长'),
    ]

    for pattern, metric_name in pattern_metric_rules:
        events = c.execute("SELECT id, name FROM events").fetchall()
        for event_id, event_name in events:
            if re.match(pattern, event_name):
                try:
                    c.execute('INSERT OR IGNORE INTO event_metric_refs (event_id, metric_name) VALUES (?, ?)',
                              (event_id, metric_name))
                    linked += 1
                except sqlite3.IntegrityError:
                    pass

    conn.commit()
    count = c.execute('SELECT COUNT(*) FROM event_metric_refs').fetchone()[0]
    print(f'  事件-指标关联: {count} 条')


def clean_dirty_data(conn):
    """清理脏数据"""
    c = conn.cursor()

    dirty_event_names = [
        'event name', 'event name\n（English）', '事件名（英文）',
        'event name（English）', 'ssion',
    ]
    for name in dirty_event_names:
        c.execute("DELETE FROM events WHERE name=?", (name,))

    dirty_field_names = [
        'key name', 'key name\n（English）', '属性名（英文）',
        '属性名(英文)', 'key name（English）',
    ]
    for name in dirty_field_names:
        c.execute("DELETE FROM event_fields WHERE name=?", (name,))

    c.execute("DELETE FROM events WHERE name_cn='' AND report_when='' AND category='engineering'")

    conn.commit()

    events_count = c.execute('SELECT COUNT(*) FROM events').fetchone()[0]
    fields_count = c.execute('SELECT COUNT(*) FROM event_fields').fetchone()[0]
    print(f'  清理后: 事件={events_count}, 字段={fields_count}')


def verify(conn):
    c = conn.cursor()
    print('\n===== 数据验证 =====')

    tables = ['metadata', 'categories', 'sdk_fields', 'common_field_groups', 'common_fields',
              'pages', 'events', 'event_fields', 'event_common_field_refs',
              'event_page_refs', 'event_metric_refs']
    for t in tables:
        try:
            count = c.execute(f'SELECT COUNT(*) FROM {t}').fetchone()[0]
            print(f'  {t}: {count} 条')
        except sqlite3.OperationalError:
            print(f'  {t}: 不存在')

    print('\n  事件分类统计:')
    for row in c.execute('SELECT category, COUNT(*) FROM events GROUP BY category ORDER BY COUNT(*) DESC'):
        print(f'    {row[0]}: {row[1]}')

    print('\n  实验ID字段:')
    for row in c.execute("SELECT name, name_cn, type FROM common_fields WHERE name IN ('eid', 'exp_id', 'new_eid')"):
        print(f'    {row[0]} ({row[1]}): {row[2]}')


def main():
    print(f'读取Excel: {EXCEL_PATH}')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        print(f'删除旧数据库: {DB_PATH}')

    conn = sqlite3.connect(DB_PATH)
    print('创建数据库schema...')
    create_schema(conn)

    print('导入元数据...')
    import_metadata(conn)

    print('导入分类定义...')
    import_categories(conn)

    print('导入SDK系统属性...')
    import_sdk_fields(conn, wb)

    print('导入公共字段...')
    import_common_fields(conn, wb)

    print('导入页面定义...')
    import_pages(conn, wb)

    print('导入事件...')
    import_events(conn, wb)

    print('清理脏数据...')
    clean_dirty_data(conn)

    print('自动关联事件与页面...')
    auto_link_pages(conn)

    print('自动关联事件与指标...')
    auto_link_metrics(conn)

    verify(conn)

    conn.close()
    print(f'\n数据库已生成: {DB_PATH}')


if __name__ == '__main__':
    main()
