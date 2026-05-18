#!/usr/bin/env python3
"""
将埋点事件数据从 Excel 导入 SQLite 数据库。
数据库设计为 5 张表，支持灵活的关联查询。

表结构：
  1. common_field_groups - 公共字段组（预置参数/公共参数/信息流公共参数）
  2. common_fields - 公共字段定义
  3. events - 事件定义
  4. event_fields - 事件特有字段
  5. event_relations - 事件与指标/页面的关联关系

查询示例：
  - 按指标查事件: SELECT e.* FROM events e JOIN event_relations r ON e.id=r.event_id WHERE r.relation_type='metric' AND r.relation_name='浏览器信息流人均时长'
  - 按页面查事件: SELECT e.* FROM events e JOIN event_relations r ON e.id=r.event_id WHERE r.relation_type='page' AND r.relation_name='首页/推荐页'
  - 按字段查事件: SELECT DISTINCT e.* FROM events e JOIN event_fields f ON e.id=f.event_id WHERE f.name='duration'
  - 查事件的完整字段: SELECT * FROM event_fields WHERE event_id=? UNION SELECT cf.* FROM common_fields cf JOIN event_common_field_refs r ON cf.id=r.common_field_id WHERE r.event_id=?
"""

import sqlite3
import os
import openpyxl
import re

DB_PATH = os.path.join(os.path.dirname(__file__), "event_tracking.db")
XLSX_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "..",
    "Downloads",
    "内容中心onetrack埋点(三期)  带沉浸式.xlsx",
)

XLSX_PATH_RESOLVED = os.path.expanduser(
    "~/Downloads/内容中心onetrack埋点(三期)  带沉浸式.xlsx"
)


def create_schema(cur):
    cur.executescript("""
        DROP TABLE IF EXISTS event_common_field_refs;
        DROP TABLE IF EXISTS event_relations;
        DROP TABLE IF EXISTS event_fields;
        DROP TABLE IF EXISTS events;
        DROP TABLE IF EXISTS common_fields;
        DROP TABLE IF EXISTS common_field_groups;

        CREATE TABLE common_field_groups (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT NOT NULL UNIQUE,
            description TEXT,
            scope       TEXT
        );

        CREATE TABLE common_fields (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id    INTEGER NOT NULL,
            name        TEXT NOT NULL,
            name_cn     TEXT,
            type        TEXT NOT NULL,
            description TEXT,
            value_note  TEXT,
            FOREIGN KEY (group_id) REFERENCES common_field_groups(id)
        );

        CREATE TABLE events (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT NOT NULL UNIQUE,
            name_cn     TEXT NOT NULL,
            category    TEXT,
            report_when TEXT,
            description TEXT,
            note        TEXT
        );

        CREATE TABLE event_fields (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            event_id    INTEGER NOT NULL,
            name        TEXT NOT NULL,
            name_cn     TEXT,
            type        TEXT NOT NULL,
            description TEXT,
            value_note  TEXT,
            scope_note  TEXT,
            sort_order  INTEGER DEFAULT 0,
            FOREIGN KEY (event_id) REFERENCES events(id)
        );

        CREATE TABLE event_relations (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            event_id      INTEGER NOT NULL,
            relation_type TEXT NOT NULL,
            relation_name TEXT NOT NULL,
            FOREIGN KEY (event_id) REFERENCES events(id)
        );

        CREATE TABLE event_common_field_refs (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            event_id      INTEGER NOT NULL,
            group_id      INTEGER NOT NULL,
            FOREIGN KEY (event_id) REFERENCES events(id),
            FOREIGN KEY (group_id) REFERENCES common_field_groups(id)
        );

        CREATE INDEX IF NOT EXISTS idx_event_fields_name ON event_fields(name);
        CREATE INDEX IF NOT EXISTS idx_event_relations_type_name ON event_relations(relation_type, relation_name);
        CREATE INDEX IF NOT EXISTS idx_events_category ON events(category);
        CREATE INDEX IF NOT EXISTS idx_events_name ON events(name);
        CREATE INDEX IF NOT EXISTS idx_common_fields_name ON common_fields(name);
    """)


def import_common_fields(cur, wb):
    ws = wb["common key 预置&公共属性"]
    current_group_id = None

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row or not row[0]:
            continue
        field_type = str(row[0]).strip()
        name_en = str(row[1]).strip() if row[1] else None
        name_cn = str(row[2]).strip() if row[2] else None
        ftype = str(row[3]).strip() if row[3] else None
        desc = str(row[4]).strip() if row[4] else None
        value_note = str(row[5]).strip() if row[5] else None

        if field_type in ("预置参数", "公共参数"):
            cur.execute(
                "INSERT OR IGNORE INTO common_field_groups (name, description, scope) VALUES (?, ?, ?)",
                (field_type, f"内容中心onetrack {field_type}", field_type),
            )
            cur.execute("SELECT id FROM common_field_groups WHERE name=?", (field_type,))
            current_group_id = cur.fetchone()[0]

        if current_group_id and name_en and name_en != "key name（English）":
            cur.execute(
                "INSERT INTO common_fields (group_id, name, name_cn, type, description, value_note) VALUES (?, ?, ?, ?, ?, ?)",
                (current_group_id, name_en, name_cn, ftype or "string", desc, value_note),
            )

    ws2 = wb["content key 内容通用属性 "]
    cur.execute(
        "INSERT OR IGNORE INTO common_field_groups (name, description, scope) VALUES (?, ?, ?)",
        ("信息流公共参数", "信息流内容通用属性", "content"),
    )
    cur.execute("SELECT id FROM common_field_groups WHERE name='信息流公共参数'")
    content_group_id = cur.fetchone()[0]

    for row in ws2.iter_rows(min_row=3, max_row=ws2.max_row, values_only=True):
        if not row or not row[0]:
            continue
        param_type = str(row[0]).strip()
        name_en = str(row[1]).strip() if row[1] else None
        name_cn = str(row[2]).strip() if row[2] else None
        ftype = str(row[3]).strip() if row[3] else None
        desc = str(row[4]).strip() if row[4] else None
        value_note = str(row[5]).strip() if row[5] else None

        if name_en and name_en != "key name（English）":
            cur.execute(
                "INSERT INTO common_fields (group_id, name, name_cn, type, description, value_note) VALUES (?, ?, ?, ?, ?, ?)",
                (content_group_id, name_en, name_cn, ftype or "string", desc, value_note),
            )


def parse_event_rows(ws):
    events = []
    current_event = None
    header_row_count = 2

    for row in ws.iter_rows(min_row=header_row_count + 1, max_row=ws.max_row, values_only=True):
        if not row or all(v is None for v in row):
            continue

        event_name_en = row[0]
        event_name_cn = row[1]
        report_when = row[2]
        field_name = row[3]
        field_name_cn = row[4]
        field_type = row[5]
        field_value_note = row[6]
        field_remark = row[7]

        if event_name_en and str(event_name_en).strip():
            event_name_en = str(event_name_en).strip()
            if current_event:
                events.append(current_event)
            current_event = {
                "names_en": [n.strip() for n in re.split(r"\n", event_name_en) if n.strip()],
                "names_cn": [n.strip() for n in re.split(r"\n", str(event_name_cn or "")) if n.strip()],
                "report_when": str(report_when or "").strip() if report_when else None,
                "fields": [],
                "common_key_refs": [],
            }
            if field_name and str(field_name).strip():
                fn = str(field_name).strip()
                if fn.startswith("common key") or fn == "common key\ncontent key":
                    current_event["common_key_refs"].append("公共参数")
                    if "content key" in fn:
                        current_event["common_key_refs"].append("信息流公共参数")
                else:
                    current_event["fields"].append({
                        "name": fn,
                        "name_cn": str(field_name_cn or "").strip() if field_name_cn else None,
                        "type": str(field_type or "string").strip() if field_type else "string",
                        "description": str(field_value_note or "").strip() if field_value_note else None,
                        "scope_note": str(field_remark or "").strip() if field_remark else None,
                    })
        else:
            if current_event:
                if field_name and str(field_name).strip():
                    fn = str(field_name).strip()
                    if fn.startswith("common key") or fn == "common key\ncontent key":
                        current_event["common_key_refs"].append("公共参数")
                        if "content key" in fn:
                            current_event["common_key_refs"].append("信息流公共参数")
                    else:
                        current_event["fields"].append({
                            "name": fn,
                            "name_cn": str(field_name_cn or "").strip() if field_name_cn else None,
                            "type": str(field_type or "string").strip() if field_type else "string",
                            "description": str(field_value_note or "").strip() if field_value_note else None,
                            "scope_note": str(field_remark or "").strip() if field_remark else None,
                        })

    if current_event:
        events.append(current_event)
    return events


CATEGORY_MAP = {
    "app通用": "app",
    "content内容相关": "content",
    "短剧": "skit",
    "抖音&穿山甲直播间": "livestream",
    "桌面上划入口拉新拉活": "growth",
    "operating运营位": "operation",
    "me我的页面": "me",
    "ad商业化": "ad",
    "search搜索": "search",
    "异常事件捕获": "exception",
    "tab标签": "tab",
    "【8.0】激励体系": "incentive",
}


def import_events(cur, wb):
    sheets_to_import = [
        "app通用",
        "content内容相关",
        "短剧",
        "抖音&穿山甲直播间",
        "桌面上划入口拉新拉活",
        "operating运营位",
        "me我的页面",
        "ad商业化",
        "search搜索",
        "异常事件捕获",
        "tab标签",
        "【8.0】激励体系",
    ]

    for sheet_name in sheets_to_import:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        category = CATEGORY_MAP.get(sheet_name, sheet_name)
        parsed_events = parse_event_rows(ws)

        for ev in parsed_events:
            names_en = ev["names_en"]
            names_cn = ev["names_cn"]

            for i, name_en in enumerate(names_en):
                name_cn = names_cn[i] if i < len(names_cn) else name_en

                cur.execute(
                    "INSERT OR IGNORE INTO events (name, name_cn, category, report_when, description, note) VALUES (?, ?, ?, ?, ?, ?)",
                    (
                        name_en,
                        name_cn,
                        category,
                        ev["report_when"],
                        name_cn,
                        None,
                    ),
                )
                cur.execute("SELECT id FROM events WHERE name=?", (name_en,))
                event_id = cur.fetchone()[0]

                for field in ev["fields"]:
                    cur.execute(
                        "INSERT INTO event_fields (event_id, name, name_cn, type, description, value_note, scope_note) VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (
                            event_id,
                            field["name"],
                            field["name_cn"],
                            field["type"],
                            field["description"],
                            None,
                            field["scope_note"],
                        ),
                    )

                for ref in ev["common_key_refs"]:
                    cur.execute("SELECT id FROM common_field_groups WHERE name=?", (ref,))
                    row = cur.fetchone()
                    if row:
                        cur.execute(
                            "INSERT OR IGNORE INTO event_common_field_refs (event_id, group_id) VALUES (?, ?)",
                            (event_id, row[0]),
                        )


def import_relations_from_yaml(cur):
    from yaml import safe_load

    yaml_path = os.path.join(os.path.dirname(__file__), "event_tracking.yaml")
    if not os.path.exists(yaml_path):
        print(f"YAML file not found: {yaml_path}, skipping relation import")
        return

    with open(yaml_path, "r", encoding="utf-8") as f:
        try:
            data = safe_load(f)
        except Exception as e:
            print(f"YAML parse error: {e}, skipping relation import")
            return

    if not data or "events" not in data:
        print("No events found in YAML, skipping relation import")
        return

    for ev in data["events"]:
        event_name = ev.get("name")
        if not event_name:
            continue

        cur.execute("SELECT id FROM events WHERE name=?", (event_name,))
        row = cur.fetchone()
        if not row:
            continue
        event_id = row[0]

        for metric in ev.get("related_metrics", []):
            if metric:
                cur.execute(
                    "INSERT INTO event_relations (event_id, relation_type, relation_name) VALUES (?, ?, ?)",
                    (event_id, "metric", metric),
                )

        for page in ev.get("related_pages", []):
            if page:
                cur.execute(
                    "INSERT INTO event_relations (event_id, relation_type, relation_name) VALUES (?, ?, ?)",
                    (event_id, "page", page),
                )


def verify(cur):
    print("\n=== 数据库验证 ===")
    cur.execute("SELECT COUNT(*) FROM common_field_groups")
    print(f"公共字段组: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM common_fields")
    print(f"公共字段: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM events")
    print(f"事件: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM event_fields")
    print(f"事件字段: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM event_relations")
    print(f"关联关系: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM event_common_field_refs")
    print(f"公共字段引用: {cur.fetchone()[0]}")

    cur.execute("SELECT category, COUNT(*) FROM events GROUP BY category ORDER BY COUNT(*) DESC")
    print("\n事件分类统计:")
    for row in cur.fetchall():
        print(f"  {row[0]}: {row[1]}")

    print("\n=== 查询示例 ===")
    cur.execute("""
        SELECT e.name, e.name_cn FROM events e
        JOIN event_relations r ON e.id=r.event_id
        WHERE r.relation_type='metric' AND r.relation_name='浏览器信息流人均时长'
    """)
    print("与'浏览器信息流人均时长'指标关联的事件:")
    for row in cur.fetchall():
        print(f"  {row[0]} ({row[1]})")

    cur.execute("""
        SELECT e.name, e.name_cn FROM events e
        JOIN event_relations r ON e.id=r.event_id
        WHERE r.relation_type='page' AND r.relation_name='沉浸式详情页'
    """)
    print("\n发生在'沉浸式详情页'的事件:")
    for row in cur.fetchall():
        print(f"  {row[0]} ({row[1]})")

    cur.execute("""
        SELECT DISTINCT e.name, e.name_cn FROM events e
        JOIN event_fields f ON e.id=f.event_id
        WHERE f.name='duration'
    """)
    print("\n包含'duration'字段的事件:")
    for row in cur.fetchall():
        print(f"  {row[0]} ({row[1]})")


def main():
    xlsx_path = XLSX_PATH_RESOLVED
    if not os.path.exists(xlsx_path):
        xlsx_path = XLSX_PATH
    if not os.path.exists(xlsx_path):
        print(f"Excel file not found: {xlsx_path}")
        return

    print(f"Reading Excel: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        print(f"Removed old database: {DB_PATH}")

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    print("Creating schema...")
    create_schema(cur)

    print("Importing common fields...")
    import_common_fields(cur, wb)

    print("Importing events...")
    import_events(cur, wb)

    print("Importing relations from YAML...")
    import_relations_from_yaml(cur)

    conn.commit()
    verify(cur)
    conn.close()
    print(f"\nDatabase created: {DB_PATH}")


if __name__ == "__main__":
    main()
